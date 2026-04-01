from io import BytesIO

from openpyxl import load_workbook

from trip_operator.reporting import build_trip_report_filename, build_trip_report_workbook


def test_build_trip_report_workbook_creates_four_sheets():
    report = {
        "status": {
            "trip": {
                "trip_id": "las-vegas-20260403-demo01",
                "destination": "Las Vegas",
                "home_city": "Boise",
                "start_date": "2026-04-03",
                "end_date": "2026-04-07",
                "travelers": 1,
                "budget_usd": 700.0,
                "estimated_total_usd": 1160.31,
                "notes": "Keep it cheap.",
                "calendar_path": "/tmp/demo.ics",
                "metadata": {
                    "destination": {
                        "requested": "Las Vegas",
                        "resolved_name": "Las Vegas, Clark County, Nevada, United States",
                    },
                    "home_city": {"display_name": "Boise, Idaho, United States"},
                    "dates": {"days": 5, "nights": 4},
                    "budget": {"travelers": 1},
                    "cost_profile": {"tier": "standard", "currency_code": "USD"},
                    "estimated_costs": {
                        "transport_usd": 178.31,
                        "lodging_usd": 432.0,
                        "food_usd": 260.0,
                        "local_transport_usd": 90.0,
                        "activities_usd": 200.0,
                        "total_usd": 1160.31,
                        "transport_details": {"mode": "budget flight"},
                        "destination_currency_code": "USD",
                        "total_in_destination_currency": 1160.31,
                    },
                    "assessment": {
                        "budget_gap_usd": -460.31,
                        "recommended_budget_usd": 1276.34,
                        "within_budget": False,
                    },
                    "tips": ["Reduce hotel nights first."],
                    "service_sources": {
                        "geocoding": "nominatim",
                        "weather": "open-meteo",
                        "exchange_rate": "identity",
                    },
                },
            },
            "expenses": [
                {
                    "id": 1,
                    "expense_date": "2026-04-03",
                    "category": "food",
                    "amount_usd": 24.5,
                    "vendor": "Cafe",
                    "notes": "Lunch",
                    "created_at": "2026-03-31T12:00:00+00:00",
                }
            ],
            "expense_count": 1,
            "spent_total_usd": 24.5,
            "remaining_budget_usd": 675.5,
            "pace_delta_usd": -10.0,
            "on_track": True,
            "category_breakdown": [{"category": "food", "total_usd": 24.5}],
            "budget_vs_estimate_usd": -460.31,
        },
        "weather": {
            "destination": "Las Vegas, Clark County, Nevada, United States",
            "weather": {
                "summary": "Clear with average highs near 27.5C and lows near 19.0C.",
                "average_high_c": 27.5,
                "average_low_c": 19.0,
                "rain_risk_days": 0,
            },
        },
        "reminders": {
            "daily_plan": [
                {
                    "date": "2026-04-03",
                    "day_number": 1,
                    "theme": "Arrival and orientation",
                    "focus": "Keep day one light.",
                    "morning": "Check in.",
                    "afternoon": "Walk the Strip.",
                    "evening": "Budget dinner.",
                    "spend_target_usd": 95.0,
                }
            ],
            "events": [
                {
                    "date": "2026-03-31",
                    "title": "Check weather and pack",
                    "description": "Refresh weather for Las Vegas and finalize the packing list.",
                },
                {
                    "date": "2026-04-03",
                    "title": "Departure day",
                    "description": "Trip starts today.",
                },
            ]
        },
    }

    workbook_bytes = build_trip_report_workbook(report)
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert workbook.sheetnames == ["Travel MCP", "Budget MCP", "Calendar MCP", "Trip Detail"]

    travel_values = [value for row in workbook["Travel MCP"].iter_rows(values_only=True) for value in row if value is not None]
    budget_values = [value for row in workbook["Budget MCP"].iter_rows(values_only=True) for value in row if value is not None]
    calendar_values = [value for row in workbook["Calendar MCP"].iter_rows(values_only=True) for value in row if value is not None]
    detail_values = [value for row in workbook["Trip Detail"].iter_rows(values_only=True) for value in row if value is not None]

    assert "Las Vegas, Clark County, Nevada, United States" in travel_values
    assert "budget flight" in travel_values
    assert "food" in budget_values
    assert "Check weather and pack" in calendar_values
    assert "Arrival and orientation" in calendar_values
    assert "report.status.trip.trip_id" in detail_values
    assert "las-vegas-20260403-demo01" in detail_values


def test_build_trip_report_filename_sanitizes_characters():
    assert build_trip_report_filename("demo trip/2026") == "demo-trip-2026.xlsx"
