from __future__ import annotations

import json
from datetime import UTC, date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

TITLE_FILL = PatternFill("solid", fgColor="8E3E1E")
SECTION_FILL = PatternFill("solid", fgColor="F0D2BF")
HEADER_FILL = PatternFill("solid", fgColor="FAEEE6")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
SECTION_FONT = Font(color="5B240C", bold=True)
HEADER_FONT = Font(color="3C261B", bold=True)
LABEL_FONT = Font(color="5D5147", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D6C2B4"),
    right=Side(style="thin", color="D6C2B4"),
    top=Side(style="thin", color="D6C2B4"),
    bottom=Side(style="thin", color="D6C2B4"),
)


def build_trip_report_workbook(report: dict[str, Any]) -> bytes:
    workbook = Workbook()
    travel_sheet = workbook.active
    travel_sheet.title = "Travel MCP"
    budget_sheet = workbook.create_sheet("Budget MCP")
    calendar_sheet = workbook.create_sheet("Calendar MCP")
    detail_sheet = workbook.create_sheet("Trip Detail")

    _populate_travel_sheet(travel_sheet, report)
    _populate_budget_sheet(budget_sheet, report)
    _populate_calendar_sheet(calendar_sheet, report)
    _populate_detail_sheet(detail_sheet, report)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_trip_report_filename(trip_id: str) -> str:
    safe_trip_id = "".join(character if character.isalnum() or character in {"-", "_"} else "-" for character in trip_id)
    return f"{safe_trip_id or 'trip-report'}.xlsx"


def _populate_travel_sheet(worksheet: Worksheet, report: dict[str, Any]) -> None:
    status = report.get("status") or {}
    trip = status.get("trip") or {}
    snapshot = trip.get("metadata") or {}
    destination = snapshot.get("destination") or {}
    dates = snapshot.get("dates") or {}
    budget = snapshot.get("budget") or {}
    weather = (report.get("weather") or {}).get("weather") or snapshot.get("weather") or {}
    costs = snapshot.get("estimated_costs") or {}
    profile = snapshot.get("cost_profile") or {}
    assessment = snapshot.get("assessment") or {}
    sources = snapshot.get("service_sources") or {}
    home_city = snapshot.get("home_city") or {}

    row = _write_sheet_title(worksheet, "Travel MCP")
    row = _write_key_value_section(
        worksheet,
        row,
        "Resolved Trip Context",
        [
            ("Trip ID", trip.get("trip_id")),
            ("Requested destination", destination.get("requested") or trip.get("destination")),
            ("Resolved destination", destination.get("resolved_name") or (report.get("weather") or {}).get("destination")),
            ("Home city", home_city.get("display_name") or trip.get("home_city")),
            ("Start date", trip.get("start_date") or dates.get("start_date")),
            ("End date", trip.get("end_date") or dates.get("end_date")),
            ("Days", dates.get("days")),
            ("Nights", dates.get("nights")),
            ("Travelers", budget.get("travelers") or trip.get("travelers")),
        ],
    )
    row = _write_key_value_section(
        worksheet,
        row,
        "Live Travel Intelligence",
        [
            ("Weather summary", weather.get("summary")),
            ("Average high (C)", weather.get("average_high_c")),
            ("Average low (C)", weather.get("average_low_c")),
            ("Rain risk days", weather.get("rain_risk_days")),
            ("Cost tier", profile.get("tier")),
            ("Currency code", profile.get("currency_code")),
            ("Destination total", _format_currency(costs.get("total_in_destination_currency"), costs.get("destination_currency_code"))),
            ("Budget gap", _format_currency(assessment.get("budget_gap_usd"), "USD")),
            ("Recommended budget", _format_currency(assessment.get("recommended_budget_usd"), "USD")),
            ("Geocoding source", sources.get("geocoding")),
            ("Weather source", sources.get("weather")),
            ("Exchange source", sources.get("exchange_rate")),
        ],
    )
    _write_table_section(
        worksheet,
        row,
        "Estimated Costs",
        ["Category", "Amount", "Details"],
        [
            ["Transport", _format_currency(costs.get("transport_usd"), "USD"), (costs.get("transport_details") or {}).get("mode")],
            ["Lodging", _format_currency(costs.get("lodging_usd"), "USD"), ""],
            ["Food", _format_currency(costs.get("food_usd"), "USD"), ""],
            ["Local transport", _format_currency(costs.get("local_transport_usd"), "USD"), ""],
            ["Activities", _format_currency(costs.get("activities_usd"), "USD"), ""],
            ["Total", _format_currency(costs.get("total_usd"), "USD"), "Projected spend"],
        ],
        footer_items=[("Tips", snapshot.get("tips") or [])],
    )
    _finalize_sheet(worksheet)


def _populate_budget_sheet(worksheet: Worksheet, report: dict[str, Any]) -> None:
    status = report.get("status") or {}
    trip = status.get("trip") or {}
    category_breakdown = status.get("category_breakdown") or []
    expenses = status.get("expenses") or []

    row = _write_sheet_title(worksheet, "Budget MCP")
    row = _write_key_value_section(
        worksheet,
        row,
        "Budget Summary",
        [
            ("Trip ID", trip.get("trip_id")),
            ("Destination", trip.get("destination")),
            ("Budget", _format_currency(trip.get("budget_usd"), "USD")),
            ("Estimated total", _format_currency(trip.get("estimated_total_usd"), "USD")),
            ("Spent so far", _format_currency(status.get("spent_total_usd"), "USD")),
            ("Remaining budget", _format_currency(status.get("remaining_budget_usd"), "USD")),
            ("Pace delta", _format_currency(status.get("pace_delta_usd"), "USD")),
            ("On track", "Yes" if status.get("on_track") else "No"),
            ("Calendar path", trip.get("calendar_path")),
            ("Notes", trip.get("notes")),
        ],
    )
    row = _write_table_section(
        worksheet,
        row,
        "Category Breakdown",
        ["Category", "Total (USD)"],
        [[item.get("category"), item.get("total_usd")] for item in category_breakdown],
    )
    _write_table_section(
        worksheet,
        row,
        "Expenses",
        ["ID", "Date", "Category", "Amount (USD)", "Vendor", "Notes", "Created at"],
        [
            [
                item.get("id"),
                item.get("expense_date"),
                item.get("category"),
                item.get("amount_usd"),
                item.get("vendor"),
                item.get("notes"),
                item.get("created_at"),
            ]
            for item in expenses
        ],
    )
    _finalize_sheet(worksheet)


def _populate_calendar_sheet(worksheet: Worksheet, report: dict[str, Any]) -> None:
    status = report.get("status") or {}
    trip = status.get("trip") or {}
    reminder_payload = report.get("reminders") or {}
    reminders = reminder_payload.get("events") or []
    daily_plan = reminder_payload.get("daily_plan") or []

    row = _write_sheet_title(worksheet, "Calendar MCP")
    row = _write_key_value_section(
        worksheet,
        row,
        "Calendar Summary",
        [
            ("Trip ID", trip.get("trip_id")),
            ("Destination", trip.get("destination")),
            ("Calendar path", trip.get("calendar_path")),
            ("Reminder count", len(reminders)),
            ("Planned day count", len(daily_plan)),
            ("Trip start", trip.get("start_date")),
            ("Trip end", trip.get("end_date")),
        ],
    )
    row = _write_table_section(
        worksheet,
        row,
        "Suggested Day Plan",
        ["Date", "Day", "Theme", "Focus", "Morning", "Afternoon", "Evening", "Spend target (USD)"],
        [
            [
                day.get("date"),
                day.get("day_number"),
                day.get("theme"),
                day.get("focus"),
                day.get("morning"),
                day.get("afternoon"),
                day.get("evening"),
                day.get("spend_target_usd"),
            ]
            for day in daily_plan
        ],
    )
    _write_table_section(
        worksheet,
        row,
        "Reminder Events",
        ["Date", "Title", "Description"],
        [[event.get("date"), event.get("title"), event.get("description")] for event in reminders],
    )
    _finalize_sheet(worksheet)


def _populate_detail_sheet(worksheet: Worksheet, report: dict[str, Any]) -> None:
    detail_rows: list[list[Any]] = []
    flattened: list[tuple[str, Any]] = [
        ("generated_at_utc", datetime.now(UTC).isoformat(timespec="seconds")),
    ]
    _flatten_payload("report", report, flattened)
    for path, value in flattened:
        detail_rows.append([path, _display_value(value)])

    row = _write_sheet_title(worksheet, "Trip Detail")
    _write_table_section(
        worksheet,
        row,
        "All Trip Details",
        ["Path", "Value"],
        detail_rows,
    )
    _finalize_sheet(worksheet)


def _write_sheet_title(worksheet: Worksheet, title: str) -> int:
    worksheet.merge_cells("A1:D1")
    cell = worksheet["A1"]
    cell.value = title
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 24
    return 3


def _write_key_value_section(
    worksheet: Worksheet,
    start_row: int,
    title: str,
    items: list[tuple[str, Any]],
) -> int:
    section_cell = worksheet.cell(start_row, 1, title)
    section_cell.fill = SECTION_FILL
    section_cell.font = SECTION_FONT
    section_cell.border = THIN_BORDER
    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)

    row = start_row + 1
    for label, value in items:
        worksheet.cell(row, 1, label).font = LABEL_FONT
        worksheet.cell(row, 1).border = THIN_BORDER
        worksheet.cell(row, 2, _display_value(value)).border = THIN_BORDER
        worksheet.cell(row, 2).alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
    return row + 1


def _write_table_section(
    worksheet: Worksheet,
    start_row: int,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    footer_items: list[tuple[str, Any]] | None = None,
) -> int:
    section_cell = worksheet.cell(start_row, 1, title)
    section_cell.fill = SECTION_FILL
    section_cell.font = SECTION_FONT
    section_cell.border = THIN_BORDER
    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max(4, len(headers)))

    header_row = start_row + 1
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(header_row, column_index, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    body_rows = rows or [["No data available", *[""] * (len(headers) - 1)]]
    row = header_row + 1
    for body in body_rows:
        for column_index, value in enumerate(body, start=1):
            cell = worksheet.cell(row, column_index, _display_value(value))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    for label, value in footer_items or []:
        row += 1
        worksheet.cell(row, 1, label).font = LABEL_FONT
        worksheet.cell(row, 1).border = THIN_BORDER
        worksheet.cell(row, 2, _display_value(value)).border = THIN_BORDER
        worksheet.cell(row, 2).alignment = Alignment(vertical="top", wrap_text=True)
    return row + 2


def _flatten_payload(prefix: str, value: Any, output: list[tuple[str, Any]]) -> None:
    if isinstance(value, dict):
        if not value:
            output.append((prefix, "{}"))
            return
        for key, nested_value in value.items():
            next_prefix = f"{prefix}.{key}" if prefix else key
            _flatten_payload(next_prefix, nested_value, output)
        return
    if isinstance(value, list):
        if not value:
            output.append((prefix, "[]"))
            return
        for index, nested_value in enumerate(value):
            _flatten_payload(f"{prefix}[{index}]", nested_value, output)
        return
    output.append((prefix, value))


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, list):
        return "; ".join(_display_value(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, sort_keys=True)
    return str(value)


def _format_currency(value: Any, currency_code: str | None) -> str:
    if value is None or value == "":
        return ""
    if currency_code and currency_code != "USD":
        return f"{value} {currency_code}"
    return f"${float(value):.2f}"


def _finalize_sheet(worksheet: Worksheet) -> None:
    worksheet.freeze_panes = "A3"
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            column_letter = cell.column_letter
            current_width = worksheet.column_dimensions[column_letter].width or 12
            worksheet.column_dimensions[column_letter].width = min(max(current_width, length + 2), 48)
